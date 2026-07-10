#!/usr/bin/env python3
"""
update_workbook.py
------------------
Applies fresh account balances (and, optionally, "actuals rolled in" month
markers) to the workbook:

    11-finance/Clent-Jewell-Finance-FY27.xlsx

Dependency: openpyxl only (`pip install openpyxl` if not already present).

USAGE
-----
    python3 update_workbook.py --set 'nab_offset=52922.84@2026-07-07' \
                                --set 'hub24_super=406323.88@2026-07-06' \
                                [--set 'key=value@date' ...] \
                                [--workbook PATH] [--dry-run]

    python3 update_workbook.py --mark-actual 2026-06 [--workbook PATH]

Each --set argument has the form:

    <state.json account key>=<numeric value>@<ISO as-at date YYYY-MM-DD>

The key MUST already exist in CELL_MAP below (which is keyed the same way
as the "accounts" array in state.json, plus a few workbook-only entries
for the Summary tab and Cashflow opening-balance cells). Multiple --set
flags may be supplied in a single invocation.

--mark-actual YYYY-MM converts a forecast month column on the Cashflow tab
into an "actual" column (green header) once that month has closed and its
forecast values have been manually/programmatically replaced with the real
Xero P&L figures. This script does NOT pull the P&L numbers itself -- see
WEEKLY-UPDATE.md step 4 ("roll actuals") for the human/agent procedure.

WHAT THIS SCRIPT DOES NOT DO
-----------------------------
- It does not touch state.json. Updating state.json (balances, as_at
  dates, last_run, sheet_url) is a distinct step in the weekly runbook
  (WEEKLY-UPDATE.md step 7) that happens AFTER the workbook has been
  re-uploaded to Google Drive, so the two are kept deliberately separate
  here.
- It does not reconcile or validate the result -- always run
  verify_workbook.py after this script.

NOTE: macq_cma_jig + macq_accel_jig share NW cell B19 — write their SUM. nab_biz_trans/maxxim/petty are Xero-domain (Jewell Group equity B22) — state.json only.
CELL_MAP below is scaffolded with placeholder cell addresses. Before this
script can be used for real, open the actual workbook and fill in every
"TODO" cell reference with the correct sheet name + cell address.
"""

import argparse
import datetime as _dt
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    sys.exit(
        "ERROR: openpyxl is required. Install it with: pip install openpyxl"
    )

DEFAULT_WORKBOOK = Path(__file__).parent / "Clent-Jewell-Finance-FY27.xlsx"

# ---------------------------------------------------------------------------
# CELL_MAP
#
# One entry per state.json account key, pointing at the workbook cell that
# holds its balance, and the cell that holds its "as at" date. Sheet names
# and cell addresses are PLACEHOLDERS -- fill these in once the real
# workbook exists (orchestrator step).
#
# Also includes a handful of non-account, workbook-level cells:
#   - cashflow_opening_cash: the Cashflow tab's current opening-cash cell,
#     which should track the sum (or a specific subset) of live bank
#     balances so the closing-cash chain starts from reality each week.
#   - summary_last_updated / summary_version: the Situation Summary tab's
#     "as at" stamp and version counter.
# ---------------------------------------------------------------------------
CELL_MAP = {
    # --- NAB accounts ---
    "nab_offset":        {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B6", "as_at_cell": "C6"},
    "nab_personal":      {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B7", "as_at_cell": "C7"},
    "nab_child_support": {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B8", "as_at_cell": "C8"},
    "nab_float":         {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B10", "as_at_cell": "C10"},
    "nab_biz_trans":     {"sheet": None, "value_cell": None, "as_at_cell": None},  # inside Jewell Group equity (Xero-sourced) — update state.json only
    "nab_maxxim":        {"sheet": None, "value_cell": None, "as_at_cell": None},  # inside Jewell Group equity (Xero-sourced) — update state.json only
    "nab_petty_cash":    {"sheet": None, "value_cell": None, "as_at_cell": None},  # inside Jewell Group equity (Xero-sourced) — update state.json only
    "nab_biz_cc":        {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B29", "as_at_cell": "C29"},
    "nab_home_loan":     {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B26", "as_at_cell": "C26"},
    # --- Macquarie accounts ---
    "macq_cma_jig":      {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B19", "as_at_cell": "C19"},
    "macq_accel_jig":    {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B19", "as_at_cell": "C19"},
    # --- Hub24 accounts ---
    "hub24_jig":         {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B18", "as_at_cell": "C18"},
    "hub24_super":       {"sheet": "Net Worth 7 Jul 2026", "value_cell": "B15", "as_at_cell": "C15"},
    # --- Workbook-level, non-account cells ---
    "cashflow_opening_cash": {"sheet": "Cashflow Budget FY27", "value_cell": "B5", "as_at_cell": "P5"},
    "summary_last_updated":  {"sheet": "Summary — Situation", "value_cell": "A2", "as_at_cell": None},
    "summary_version":       {"sheet": "Summary — Situation", "value_cell": "A2", "as_at_cell": None},
}

# ---------------------------------------------------------------------------
# MONTH_COLUMN_MAP
#
# Maps "YYYY-MM" -> the Cashflow tab column letter holding that month, for
# use by --mark-actual. TODO: fill in once the workbook's month layout is
# known (likely one column per month across a FY27 row of headers).
# ---------------------------------------------------------------------------
MONTH_COLUMN_MAP = {
    "2026-07":"B","2026-08":"C","2026-09":"D","2026-10":"E","2026-11":"F","2026-12":"G","2027-01":"H","2027-02":"I","2027-03":"J","2027-04":"K","2027-05":"L","2027-06":"M",
    # "2026-07": "TODO",
    # "2026-08": "TODO",
    # ... etc, one per FY27 month
}
ACTUAL_HEADER_ROW = 3  # month header row on the Cashflow tab

ACTUAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def parse_set_arg(raw: str):
    """Parse 'key=value@date' into (key, float(value), date_str)."""
    if "=" not in raw or "@" not in raw:
        raise ValueError(
            f"--set value {raw!r} must look like key=value@YYYY-MM-DD"
        )
    key, rest = raw.split("=", 1)
    value_str, date_str = rest.rsplit("@", 1)
    key = key.strip()
    value = float(value_str.strip())
    date_str = date_str.strip()
    # Validate the date parses as ISO.
    _dt.date.fromisoformat(date_str)
    return key, value, date_str


def apply_set(ws_lookup, key, value, date_str, dry_run=False):
    if key not in CELL_MAP:
        raise KeyError(
            f"Unknown key {key!r} -- add it to CELL_MAP in update_workbook.py first."
        )
    mapping = CELL_MAP[key]
    for field in ("value_cell", "as_at_cell"):
        cell_ref = mapping.get(field)
        if cell_ref == "TODO":
            raise NotImplementedError(
                f"CELL_MAP['{key}']['{field}'] is still a TODO placeholder. "
                "Fill in the real cell address before running updates."
            )

    ws = ws_lookup(mapping["sheet"])
    print(f"  {key}: {mapping['sheet']}!{mapping['value_cell']} = {value}"
          f"  (as_at {mapping['as_at_cell']} = {date_str})")
    if dry_run:
        return
    ws[mapping["value_cell"]] = value
    if mapping.get("as_at_cell"):
        ws[mapping["as_at_cell"]] = date_str


def mark_month_actual(ws_lookup, month: str, dry_run=False):
    if month not in MONTH_COLUMN_MAP:
        raise KeyError(
            f"Month {month!r} not in MONTH_COLUMN_MAP -- add its column letter first."
        )
    if ACTUAL_HEADER_ROW is None:
        raise NotImplementedError(
            "ACTUAL_HEADER_ROW is not set -- fill in the Cashflow tab's month "
            "header row number first."
        )
    col = MONTH_COLUMN_MAP[month]
    cell_ref = f"{col}{ACTUAL_HEADER_ROW}"
    ws = ws_lookup("Cashflow")
    print(f"  Marking {month} ({cell_ref}) as ACTUAL (green header) -- "
          "remember to also replace that column's forecast values with Xero P&L actuals.")
    if dry_run:
        return
    ws[cell_ref].fill = ACTUAL_FILL


def bump_version_stamp(ws_lookup, dry_run=False):
    """Increment the Summary tab's version counter and stamp today's date."""
    mapping = CELL_MAP["summary_version"]
    if mapping["value_cell"] == "TODO":
        print("  (skipping version stamp bump -- CELL_MAP['summary_version'] is a TODO)")
        return
    ws = ws_lookup(mapping["sheet"])
    current = ws[mapping["value_cell"]].value
    try:
        new_version = int(current) + 1
    except (TypeError, ValueError):
        new_version = 1
    print(f"  Version stamp: {current!r} -> {new_version}")
    if not dry_run:
        ws[mapping["value_cell"]] = new_version

    last_updated_mapping = CELL_MAP["summary_last_updated"]
    if last_updated_mapping["value_cell"] != "TODO":
        today = _dt.date.today().isoformat()
        print(f"  Last-updated stamp -> {today}")
        if not dry_run:
            ws_lookup(last_updated_mapping["sheet"])[last_updated_mapping["value_cell"]] = today


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook", type=Path, default=DEFAULT_WORKBOOK,
        help=f"Path to the xlsx workbook (default: {DEFAULT_WORKBOOK})",
    )
    parser.add_argument(
        "--set", action="append", default=[], dest="sets",
        help="key=value@YYYY-MM-DD ; may be repeated.",
    )
    parser.add_argument(
        "--mark-actual", action="append", default=[], dest="mark_actuals",
        metavar="YYYY-MM",
        help="Mark a Cashflow month column as actual (green header); may be repeated.",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Print what would change without writing/saving the workbook.",
    )
    args = parser.parse_args()

    if not args.sets and not args.mark_actuals:
        parser.error("Nothing to do -- pass at least one --set or --mark-actual.")

    if not args.workbook.exists():
        sys.exit(
            f"ERROR: workbook not found at {args.workbook}. "
            "It is added by the orchestrator separately; this script only "
            "edits it once it exists."
        )

    print(f"Loading workbook: {args.workbook}")
    wb = load_workbook(args.workbook)

    def ws_lookup(sheet_name):
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Sheet {sheet_name!r} not found in workbook. Available sheets: "
                f"{wb.sheetnames}"
            )
        return wb[sheet_name]

    print("Applying balance updates:" if args.sets else "No balance updates requested.")
    for raw in args.sets:
        key, value, date_str = parse_set_arg(raw)
        apply_set(ws_lookup, key, value, date_str, dry_run=args.dry_run)

    for month in args.mark_actuals:
        mark_month_actual(ws_lookup, month, dry_run=args.dry_run)

    print("Bumping version stamp:")
    bump_version_stamp(ws_lookup, dry_run=args.dry_run)

    if args.dry_run:
        print("\nDRY RUN -- no changes written.")
        return

    wb.save(args.workbook)
    print(f"\nSaved: {args.workbook}")
    print("Next: run verify_workbook.py before re-uploading to Google Drive.")


if __name__ == "__main__":
    main()
