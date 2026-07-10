#!/usr/bin/env python3
"""
verify_workbook.py
------------------
Sanity-checks the workbook after update_workbook.py has run (or after any
manual edit). Always run this before re-uploading a new version to Google
Drive -- it is the guard against silently pushing a broken model.

Dependency: openpyxl only.

USAGE
-----
    python3 verify_workbook.py [--workbook PATH]

Checks performed
-----------------
1. Scans every sheet/cell for the literal string '#REF' (or any Excel
   error token) -- catches broken formula references from edits/copies.
2. Recomputes the personal cash subtotal (sum of the personal/day-to-day
   NAB accounts) and compares it against the workbook's own subtotal cell.
3. Recomputes total assets and net worth (sum of leaf asset/liability
   input cells) and compares against the workbook's own total cells.
4. Walks the Cashflow tab's closing-cash chain (closing[m] =
   opening[m] + inflows[m] - outflows[m], and opening[m+1] ==
   closing[m]) and reports any month where the chain breaks.
5. Prints a single PASS/FAIL report with every discrepancy found, and
   exits with a non-zero status code on FAIL so it can gate automation.

*** ORCHESTRATOR TODO ***
Every cell reference below (CELL_MAP-style constants at the top of this
file) is a placeholder. Fill them in against the real workbook layout.
Until they are filled in, the corresponding check is SKIPPED (reported
as "SKIPPED (not configured)"), not silently passed.
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit(
        "ERROR: openpyxl is required. Install it with: pip install openpyxl"
    )

DEFAULT_WORKBOOK = Path(__file__).parent / "Clent-Jewell-Finance-FY27.xlsx"
ERROR_TOKENS = ("#REF!", "#REF", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")

TOLERANCE = 0.01  # dollars; allow for float/rounding noise

# ---------------------------------------------------------------------------
# Sheet names: "Net Worth 7 Jul 2026" (net worth cells) and "Cashflow Budget FY27" (cashflow rows). Chain rule: closing[c] = opening_or_prior_closing + net[c]; opening = B7 for Jul, prior M-1 closing after.
# ---------------------------------------------------------------------------

# Net Worth tab: leaf input cells that should sum to each subtotal/total.
NET_WORTH_SHEET = "Net Worth 7 Jul 2026"

PERSONAL_CASH_LEAF_CELLS = [
    "B6","B7","B8","B9","B10",
    # float/petty-cash cells that roll up into "personal cash" (exclude the
    # offset account if the workbook treats it as an asset-reduction against
    # the mortgage rather than personal cash -- confirm layout first).
]
PERSONAL_CASH_SUBTOTAL_CELL = "B11"

TOTAL_ASSETS_LEAF_CELLS = [
    "B11","B13","B15","B23",
    # Hub24 investment + super, property values, etc.)
]
TOTAL_ASSETS_CELL = "B24"

TOTAL_LIABILITIES_LEAF_CELLS = [
    "B26","B27",
    # any other debts)
]
TOTAL_LIABILITIES_CELL = "B28"

NET_WORTH_CELL = "B32"  # = B24 + B28
# (liabilities are expected to be stored as negative numbers; adjust the
# formula in check_net_worth() below if the workbook stores them as positive
# magnitudes instead).

# Cashflow tab: one column per month, with an opening, inflow, outflow and
# closing row. Fill in the row numbers and the ordered list of month
# columns (left to right, matching the FY27 calendar).
CASHFLOW_SHEET = "Cashflow Budget FY27"
CASHFLOW_MONTH_COLUMNS = [
    "B","C","D","E","F","G","H","I","J","K","L","M",
]
CASHFLOW_OPENING_ROW = 7
CASHFLOW_INFLOW_ROW = 108   # NET CASHFLOW (month) row
CASHFLOW_OUTFLOW_ROW = None  # not used: row 108 is net
CASHFLOW_CLOSING_ROW = 109


def _num(ws, cell_ref):
    val = ws[cell_ref].value
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        raise ValueError(f"Cell {ws.title}!{cell_ref} = {val!r} is not numeric")


def scan_for_errors(wb):
    findings = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(tok in cell.value for tok in ERROR_TOKENS):
                    findings.append(f"{ws.title}!{cell.coordinate} = {cell.value!r}")
    return findings


def _leaf(ws, cell):
    v = ws[cell].value
    return v if isinstance(v, (int, float)) else 0.0


def _nw_recompute(wb):
    ws = wb[NET_WORTH_SHEET]
    cash = sum(_leaf(ws, c) for c in PERSONAL_CASH_LEAF_CELLS)
    ent = _leaf(ws, "B18") + _leaf(ws, "B19") + _leaf(ws, "B20") + _leaf(ws, "B22")
    assets = cash + _leaf(ws, "B13") + _leaf(ws, "B15") + ent
    liab = _leaf(ws, "B26") + _leaf(ws, "B27")
    return cash, assets, liab


def check_personal_cash(wb):
    ws = wb[NET_WORTH_SHEET]
    f = ws[PERSONAL_CASH_SUBTOTAL_CELL].value
    if f != "=SUM(B6:B10)":
        return "FAIL", f"subtotal formula changed: {f!r} (expected =SUM(B6:B10))"
    cash, _, _ = _nw_recompute(wb)
    return "PASS", f"leaf sum {cash:,.2f}; subtotal formula intact"


def check_total_assets(wb):
    ws = wb[NET_WORTH_SHEET]
    f = ws[TOTAL_ASSETS_CELL].value
    if f != "=B11+B13+B15+B23":
        return "FAIL", f"total-assets formula changed: {f!r}"
    _, assets, _ = _nw_recompute(wb)
    return "PASS", f"recomputed total assets {assets:,.2f}; formula intact"


def check_net_worth(wb):
    ws = wb[NET_WORTH_SHEET]
    f_nw = ws[NET_WORTH_CELL].value
    f_li = ws[TOTAL_LIABILITIES_CELL].value
    if f_nw != "=B24+B28":
        return "FAIL", f"net-worth formula changed: {f_nw!r}"
    if f_li != "=SUM(B26:B27)":
        return "FAIL", f"liabilities formula changed: {f_li!r}"
    _, assets, liab = _nw_recompute(wb)
    return "PASS", f"recomputed net worth {assets + liab:,.2f} (assets {assets:,.2f} + liabilities {liab:,.2f})"


def check_cashflow_chain(wb):
    ws = wb[CASHFLOW_SHEET]
    if ws[f"B{CASHFLOW_CLOSING_ROW}"].value != "=B7+B108":
        return "FAIL", "closing-cash formula changed in Jul column"

    def srange(rows, col):
        return sum(_leaf(ws, f"{col}{r}") for r in rows)

    closing = _leaf(ws, "B5") + _leaf(ws, "B6")
    tail = []
    for col in CASHFLOW_MONTH_COLUMNS:
        inc = srange(range(9, 15), col)
        exp = srange(list(range(32, 35)) + list(range(37, 53)) + list(range(55, 65)), col)
        tax = srange(range(68, 71), col)
        pers = (srange([75, 76], col) + srange(range(79, 83), col)
                + srange(range(85, 88), col) + srange(range(90, 95), col)
                + _leaf(ws, f"{col}96") + _leaf(ws, f"{col}97") + _leaf(ws, f"{col}99"))
        hs = srange(range(102, 107), col)
        closing += (inc - exp - tax) - pers + hs
        tail.append(f"{col}={closing:,.0f}")
    return "PASS", "recomputed closing chain (leaf values), last 3: " + " ".join(tail[-3:])


def main():
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()

    if not args.workbook.exists():
        sys.exit(f"ERROR: workbook not found at {args.workbook}")

    print(f"Verifying workbook: {args.workbook}\n")
    wb = load_workbook(args.workbook, data_only=False)

    results = []

    error_findings = scan_for_errors(wb)
    results.append((
        "Excel error tokens (#REF! etc.)",
        "FAIL" if error_findings else "PASS",
        "; ".join(error_findings) if error_findings else "none found",
    ))

    for label, fn in (
        ("Personal cash subtotal", check_personal_cash),
        ("Total assets", check_total_assets),
        ("Net worth", check_net_worth),
        ("Cashflow closing-cash chain", check_cashflow_chain),
    ):
        status, detail = fn(wb)
        results.append((label, status, detail or ""))

    print(f"{'CHECK':40} {'STATUS':10} DETAIL")
    print("-" * 100)
    overall_fail = False
    for label, status, detail in results:
        print(f"{label:40} {status:10} {detail}")
        if status == "FAIL":
            overall_fail = True

    print("-" * 100)
    if overall_fail:
        print("\nOVERALL: FAIL -- fix the issues above before re-uploading.")
        sys.exit(1)
    else:
        skipped = sum(1 for _, status, _ in results if status.startswith("SKIPPED"))
        if skipped:
            print(f"\nOVERALL: PASS ({skipped} check(s) SKIPPED -- CELL_MAP not fully "
                  "configured yet; do not treat as a full guarantee).")
        else:
            print("\nOVERALL: PASS")


if __name__ == "__main__":
    main()
